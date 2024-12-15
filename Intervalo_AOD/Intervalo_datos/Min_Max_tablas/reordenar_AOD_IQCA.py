import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Lista de archivos
archivos = [
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_diciembre_febrero_2019.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_diciembre_febrero_2020.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_diciembre_febrero_2021.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_diciembre_febrero_2024.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_junio_agosto_2019.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_junio_agosto_2020.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_junio_agosto_2021.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_junio_agosto_2024.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_marzo_mayo_2019.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_marzo_mayo_2020.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_marzo_mayo_2021.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_marzo_mayo_2024.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_septiembre_noviembre_2019.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_septiembre_noviembre_2020.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_septiembre_noviembre_2021.csv",
    r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_periodos\Resultados_septiembre_noviembre_2024.csv"
]

# Inicializar el dataframe final
df_final = pd.DataFrame()

# Procesar cada archivo
for archivo in archivos:
    # Leer el archivo CSV
    df = pd.read_csv(archivo)
    
    # Extraer el periodo de meses y el año desde el nombre del archivo
    nombre_archivo = os.path.basename(archivo)
    periodo = nombre_archivo.split('_')[1] + "_" + nombre_archivo.split('_')[2]  # Ejemplo: diciembre_febrero
    anio = nombre_archivo.split('_')[-1].split('.')[0]  # Ejemplo: 2019
    
    # Agregar el periodo de meses y el año como columnas
    df['Periodo'] = periodo
    df['Año'] = anio
    
    # Filtrar y seleccionar solo las columnas relevantes
    df = df[['Estacion', 'Categoria', 'Valor Categoria', 'AOD', 'Min', 'Max', 'Periodo', 'Año']]
    
    # Filtrar solo las filas correspondientes a la categoría 'IQCA' y AOD_moda
    df = df[(df['Categoria'] == 'IQCA') & (df['AOD'] == 'AOD_moda')]
    
    # Concatenar los datos al dataframe final
    df_final = pd.concat([df_final, df])

# Definir el orden de las columnas de 'Periodo' y 'Categoria'
orden_periodo = ['diciembre_febrero', 'marzo_mayo', 'junio_agosto', 'septiembre_noviembre']
orden_categoria = ['IQCA']
orden_valor_categoria = ['Deseable', 'Aceptable', 'Precaucion', 'Alerta', 'Alarma', 'Emergencia', 'Incorrecto']

# Crear la matriz de doble entrada (pivot)
pivot_df = df_final.melt(
    id_vars=['Año', 'Estacion', 'Periodo', 'Categoria', 'Valor Categoria'],
    value_vars=['Min', 'Max'],
    var_name='Tipo',
    value_name='Valor'
)

# Reorganizar para asegurar que Min y Max estén en filas separadas bajo cada Valor Categoria
pivot_df = pivot_df.pivot_table(
    index=['Año', 'Estacion'],
    columns=['Periodo', 'Categoria', 'Valor Categoria', 'Tipo'],
    values='Valor',
    aggfunc='first'
)

# Asegurarse de que las columnas de 'Periodo', 'Categoria' y 'Valor Categoria' sigan el orden definido
pivot_df = pivot_df.reindex(columns=pd.MultiIndex.from_product([orden_periodo, orden_categoria, orden_valor_categoria, pivot_df.columns.levels[3]]))

# Guardar el archivo Excel
ruta_salida = r"C:\Users\gisse\OneDrive\Escritorio\Repositorio\Documentos\Intervalo_AOD\Intervalo_datos\Min_Max_tablas\resultado_AOD_media_IQCA.xlsx"
pivot_df.to_excel(ruta_salida)

# Aplicar colores usando openpyxl
wb = load_workbook(ruta_salida)
ws = wb.active

# Colores para cada valor de 'Valor Categoria'
colores = {
    'Deseable': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),  # Blanco
    'Aceptable': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Verde
    'Precaucion': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Gris
    'Alerta': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Amarillo
    'Alarma': PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid"),  # Tomate
    'Emergencia': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Rojo
    'Incorrecto': PatternFill(start_color="AFEEEE", end_color="AFEEEE", fill_type="solid")  # Celeste
}

# Aplicar los colores a las columnas correspondientes (toda la columna para cada Valor Categoria)
for col in ws.columns:
    for cell in col:
        if cell.row == 1:  # Evitar colorear la fila de encabezado
            continue
        if cell.value in colores:
            # Aplicar el color a toda la columna que corresponde al valor
            col_letter = cell.column_letter
            for r in range(2, ws.max_row + 1):  # Comenzar en la fila 2 (después del encabezado)
                ws[f"{col_letter}{r}"].fill = colores[cell.value]

# Guardar el archivo con los colores aplicados
wb.save(ruta_salida)

print(f"El archivo se ha guardado en: {ruta_salida}")
